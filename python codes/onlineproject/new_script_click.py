import os
import django

import bs4 as bs

os.environ.setdefault("DJANGO_SETTINGS_MODULE", "onlineproject.settings")
django.setup()

import onlineapp.models
from onlineapp.models import *


import click
import openpyxl
import pymysql


@click.group()
def cli():
    """commands for db import/export"""
    pass



@cli.command()
def cleardata():
    """Clears all tables data"""
    MockTest1.objects.all().delete()
    Student.objects.all().delete()
    College.objects.all().delete()
    pass



@cli.command()
def createdb():
    """creates database tutdb and adds tables"""
    mysql_con = pymysql.connect(user="root", password="password", host="localhost")
    con = mysql_con.cursor()
    con.execute("CREATE DATABASE tutdb")
    mysql_con.close()

    from django.core.management import call_command
    call_command("makemigrations")
    call_command("migrate", interactive=False)

    pass



@cli.command()
def dropdb():
    """Drops database tutdb"""
    mysql_con = pymysql.connect(user="root", password="password", database="tutdb", host="localhost")
    con = mysql_con.cursor()
    con.execute("DROP DATABASE tutdb;")
    mysql_con.close()
    pass



@cli.command()
@click.argument("student_file")
@click.argument("marks_file")
def populatedb(student_file,marks_file):
    """populates db from students file and marks file"""
    #student_file="students.xlsx"
    #marks_file="marks.xlsx"

    wb_student = openpyxl.load_workbook(student_file)
    sheet_college = wb_student.get_sheet_by_name("Colleges")

    rows = sheet_college.rows
    rows.__next__()

    dict_colleges={}
    for row in rows:
        l = []
        for cell in row:
            l.append(cell.value)
        c=College(name=l[0].strip(),location=l[2].strip(),acronym=l[1].strip(),contact=l[3].strip())
        c.save()
        dict_colleges[l[1]]=c

    sheet_student= wb_student.get_sheet_by_name("Current")
    rows=sheet_student.rows
    rows.__next__()

    dict_students={}
    for row in rows:
        l = []
        for cell in row:
            l.append(cell.value)
        s=Student(name=l[0].strip().lower(),college=dict_colleges[l[1].strip()],email=l[2].strip(),db_folder="ol2016_"+l[1].strip()+"_"+l[3].strip().lower()+"_mock")
        s.save()
        dict_students[s.db_folder]=s

    sheet_student = wb_student.get_sheet_by_name("Deletions")
    rows = sheet_student.rows
    rows.__next__()

    for row in rows:
        l = []
        for cell in row:
            l.append(cell.value)
        s=Student(name=l[0].strip().lower(),college=dict_colleges[l[1].strip()],email=l[2].strip(),db_folder="ol2016_"+l[1].strip()+"_"+l[3].strip().lower()+"_mock",dropped_out=True)
        s.save()
        dict_students[s.db_folder]=s

    soup = bs.BeautifulSoup(open(marks_file), "html.parser")

    #wb = openpyxl.Workbook()
    #sheet = wb.active

    b=0
    for tr in soup.find_all("tr"):
        l2 = []
        for td in tr:
            l2.append(td.text)
        if(b==1):
            try:
                m=MockTest1(problem1=int(l2[2]),problem2=int(l2[3]),problem3=int(l2[4]),problem4=int(l2[5]),total=int(l2[6]),student=dict_students[l2[1]])
                m.save()
            except Exception as e:
                print("key not found")
        b=1

    print("Asd")







"""
    wb_marks = openpyxl.load_workbook(marks_file)
    sheet_marks = wb_marks.get_sheet_by_name("Sheet")

    rows = sheet_marks.rows
    rows.__next__()

    for row in rows:
        l = []
        for cell in row:
            l.append(cell.value)
        m=MockTest1(problem1=int(l[1].strip()),problem2=int(l[2].strip()),problem3=int(l[3].strip()),problem4=int(l[4].strip()),total=int(l[5].strip()),student=dict_students[l[0].strip()])
        m.save()

    print("Asd")
    print("asdasd")
    print("Asdsa")
"""


"""    wb_marks = openpyxl.load_workbook(marks_table)
    sheet = wb_marks.get_sheet_by_name("Sheet")
    rows = sheet.rows
    rows.__next__()
    for row in rows:
        l = []
        for cell in row:
            l.append(cell.value)
        l[0] = l[0].split("_")[2]
        if l[0].strip().lower() == "":
            continue
        con.execute(
            "insert into marks values('{}',{},{},{},{},{})".format(l[0].strip().lower(), l[1].strip(), l[2].strip(),
                                                                   l[3].strip(), l[4].strip(), l[5].strip()))

    pass
"""
#c=College(name="ChaitanyaBharathi",location="hyderabad",acronym="cbit",contact="cbit@gmail.com")
#c.save()

if __name__=="__main__":
    cli()

